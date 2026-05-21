"""Servicio del módulo de membresía.

Centraliza toda la lógica de negocio relacionada con miembros:
  - Creación y actualización de miembros
  - Alta simultánea con acceso al sistema
  - Anonimización RGPD
  - Exportación a XLSX

El resolver GraphQL (membresia_resolvers.py) solo delega en este servicio:
nunca toca la sesión directamente ni contiene lógica de dominio.
"""

from __future__ import annotations

import base64
import io
import uuid
from datetime import date
from typing import Optional, Sequence
from uuid import UUID

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..models.miembro import Miembro


# ---------------------------------------------------------------------------
# Campos mapeables directamente desde los inputs (sin transformación)
# ---------------------------------------------------------------------------

_MIEMBRO_FIELDS: tuple[str, ...] = (
    'nombre', 'apellido1', 'apellido2', 'sexo', 'fecha_nacimiento',
    'tipo_miembro_id', 'estado_id', 'tipo_documento', 'numero_documento',
    'pais_documento_id', 'pais_nacimiento_id',
    'direccion', 'codigo_postal', 'localidad',
    'provincia_id', 'pais_domicilio_id',
    'telefono', 'telefono2', 'email',
    'agrupacion_id',
    'iban', 'swift_bic', 'referencia_pago', 'forma_pago_id',
    'es_socio_honor',
    'fecha_alta', 'fecha_baja',
    'motivo_baja_id', 'motivo_baja_texto',
    'observaciones',
    'solicita_supresion_datos', 'fecha_solicitud_supresion',
    'fecha_limite_retencion', 'datos_anonimizados', 'fecha_anonimizacion',
    'activo',
    'es_voluntario', 'disponibilidad', 'horas_disponibles_semana',
    'profesion', 'nivel_estudios_id', 'motivo_reduccion_id',
    'experiencia_voluntariado', 'intereses', 'observaciones_voluntariado',
    'puede_conducir', 'vehiculo_propio', 'disponibilidad_viajar',
)

# Campos obligatorios: en actualización, un None explícito no se aplica.
_CAMPOS_REQUERIDOS: frozenset[str] = frozenset({
    'nombre', 'apellido1', 'tipo_miembro_id', 'estado_id',
})

# Campos personales que se borran al anonimizar (RGPD)
_CAMPOS_PII: tuple[str, ...] = (
    'nombre', 'apellido1', 'apellido2',
    'fecha_nacimiento', 'sexo',
    'tipo_documento', 'numero_documento',
    'direccion', 'codigo_postal', 'localidad',
    'telefono', 'telefono2', 'email',
    'iban', 'swift_bic', 'referencia_pago',
    'foto_url',
    'observaciones', 'observaciones_voluntariado',
)


class MembresiaService:
    """Orquesta operaciones sobre miembros de la organización."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    # ------------------------------------------------------------------
    # Helpers internos
    # ------------------------------------------------------------------

    async def _get(self, miembro_id: UUID) -> Miembro:
        """Recarga el miembro con todas las relaciones (selectin en el modelo)."""
        result = await self.session.execute(
            select(Miembro).where(Miembro.id == miembro_id)
        )
        miembro = result.scalar_one_or_none()
        if miembro is None:
            raise ValueError(f"No existe un miembro con id '{miembro_id}'")
        return miembro

    def _apply_fields(self, target: Miembro, data: object, *, skip_none_required: bool = False) -> None:
        """Aplica los campos de un input dataclass/strawberry sobre el modelo.

        En modo actualización (skip_none_required=True), los campos obligatorios
        con valor None se ignoran para no sobreescribir datos existentes.
        """
        for field in _MIEMBRO_FIELDS:
            value = getattr(data, field, None)
            if value is None and skip_none_required and field in _CAMPOS_REQUERIDOS:
                continue
            setattr(target, field, value)

    # ------------------------------------------------------------------
    # CRUD
    # ------------------------------------------------------------------

    async def crear(self, data: object) -> Miembro:
        """Crea un nuevo miembro a partir de un input.

        Args:
            data: Objeto con los mismos atributos que _MIEMBRO_FIELDS
                  (normalmente MiembroCreateInput).

        Returns:
            El miembro recién creado, recargado con sus relaciones.
        """
        miembro = Miembro()
        self._apply_fields(miembro, data)
        self.session.add(miembro)
        await self.session.commit()
        return await self._get(miembro.id)

    async def crear_con_acceso(
        self,
        data: object,
        email: str,
        password: str,
        tipo_vinculacion_id: Optional[UUID] = None,
        activo_usuario: bool = True,
    ) -> Miembro:
        """Crea un miembro y su usuario de acceso en una única transacción atómica.

        Delega la creación del usuario en AccesoService para no duplicar
        la lógica de hash de contraseña y validación de email único.

        Args:
            data: Input de membresía.
            email: Email del usuario de acceso.
            password: Contraseña en texto plano (se hashea en AccesoService).
            tipo_vinculacion_id: Tipo de vinculación del usuario (opcional).
            activo_usuario: Si el usuario queda activo desde el alta.

        Returns:
            El miembro creado con su relación `usuario` cargada.
        """
        # Importación local para evitar ciclo de dependencias entre módulos.
        from app.modules.acceso.services.acceso_service import AccesoService

        miembro = Miembro()
        self._apply_fields(miembro, data)
        self.session.add(miembro)
        await self.session.flush()  # necesitamos el id antes de crear el usuario

        acceso_svc = AccesoService(self.session)
        await acceso_svc.crear_usuario(
            email=email,
            password=password,
            activo=activo_usuario,
            miembro_id=miembro.id,
            tipo_vinculacion_id=tipo_vinculacion_id,
        )

        await self.session.commit()
        return await self._get(miembro.id)

    async def actualizar(self, data: object) -> Miembro:
        """Actualiza los campos de un miembro existente.

        Los campos obligatorios con valor None en el input se ignoran
        (no se sobreescriben datos existentes con None accidentalmente).

        Args:
            data: Input de actualización; debe tener un atributo `id: UUID`.

        Returns:
            El miembro actualizado, recargado con sus relaciones.
        """
        miembro = await self._get(data.id)
        self._apply_fields(miembro, data, skip_none_required=True)
        await self.session.commit()
        return await self._get(miembro.id)

    # ------------------------------------------------------------------
    # RGPD
    # ------------------------------------------------------------------

    async def anonimizar(self, miembro_id: UUID) -> Miembro:
        """Anonimiza de forma irreversible los datos personales del miembro.

        Solo puede ejecutarse sobre miembros dados de baja. El registro se
        conserva para estadística e histórico pero sin información identificativa.

        Raises:
            ValueError: si el miembro no existe, ya está anonimizado,
                        o no está dado de baja.
        """
        miembro = await self._get(miembro_id)

        if miembro.datos_anonimizados:
            raise ValueError("Los datos de este miembro ya están anonimizados.")
        if miembro.fecha_baja is None:
            raise ValueError(
                "Solo pueden anonimizarse los datos de un miembro dado de baja."
            )

        # Borrar todos los campos PII definidos de forma centralizada.
        for campo in _CAMPOS_PII:
            # nombre y apellido1 reciben un literal; el resto quedan a None.
            if campo in ('nombre', 'apellido1'):
                setattr(miembro, campo, 'Anonimizado')
            else:
                setattr(miembro, campo, None)

        miembro.datos_anonimizados = True
        miembro.fecha_anonimizacion = date.today()

        await self.session.commit()
        return await self._get(miembro_id)

    # ------------------------------------------------------------------
    # Exportación
    # ------------------------------------------------------------------

    async def exportar_xlsx(self, ids: Sequence[UUID]) -> str:
        """Exporta a XLSX los miembros indicados.

        Args:
            ids: Lista de UUIDs de miembros a exportar.

        Returns:
            Contenido del fichero XLSX codificado en base64.

        Raises:
            ValueError: si la lista de ids está vacía.
        """
        if not ids:
            raise ValueError("No hay miembros que exportar.")

        # Importación diferida: openpyxl no siempre está disponible en tests.
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        result = await self.session.execute(
            select(Miembro).where(Miembro.id.in_(ids))
        )
        miembros = list(result.scalars().all())
        miembros.sort(key=lambda m: (
            (m.apellido1 or '').lower(),
            (m.apellido2 or '').lower(),
            (m.nombre or '').lower(),
        ))

        wb = Workbook()
        ws = wb.active
        ws.title = "Miembros"

        cabecera = [
            "Nombre", "Primer apellido", "Segundo apellido",
            "Tipo", "Situación",
            "Email", "Teléfono",
            "Agrupación", "Localidad",
            "Fecha de alta", "Fecha de baja",
        ]
        ws.append(cabecera)
        for celda in ws[1]:
            celda.font = Font(bold=True)

        for m in miembros:
            ws.append([
                m.nombre or '',
                m.apellido1 or '',
                m.apellido2 or '',
                m.tipo_miembro.nombre if m.tipo_miembro else '',
                m.estado.nombre if m.estado else '',
                m.email or '',
                m.telefono or '',
                m.agrupacion.nombre if m.agrupacion else '',
                m.localidad or '',
                m.fecha_alta.isoformat() if m.fecha_alta else '',
                m.fecha_baja.isoformat() if m.fecha_baja else '',
            ])

        anchos = [16, 16, 16, 14, 14, 30, 14, 26, 20, 13, 13]
        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        return base64.b64encode(buf.getvalue()).decode()
