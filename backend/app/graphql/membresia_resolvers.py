"""Resolvers custom para membresía.

strawchemy excluye las columnas FK UUID de los inputs auto-generados.
Este módulo define MiembroCreateInput y MiembroUpdateInput completos
(con todos los FK UUID fields) y los resolvers que los usan.
"""
from __future__ import annotations

import uuid
from datetime import date
from typing import Optional, List

import strawberry
from sqlalchemy import select, or_

from app.modules.membresia.models.contacto import Contacto
from app.modules.membresia.models.participacion import Participacion, Membresia
from app.modules.membresia.models.vinculacion import Vinculacion, Socio, Voluntario
from app.modules.membresia.models.tipo_vinculacion import TipoVinculacion
from app.modules.acceso.services.acceso_service import AccesoService
from app.graphql.types_auto import ContactoType
from app.graphql.permissions import RequireTransaction, RequireAuthenticated
from app.core.events import event_bus, MiembroPerfilIncompleto


# Campos clave de perfil que ahora viven en Contacto (email/teléfono). El tipo de
# socio y el estado viven en Membresia/Socio y no se chequean aquí.
_CAMPOS_PERFIL_CLAVE: tuple[tuple[str, str], ...] = (
    ("email",    "email"),
    ("telefono", "teléfono"),
)


def _campos_perfil_faltantes(contacto) -> tuple[str, ...]:
    """Devuelve los nombres legibles de los campos clave sin rellenar."""
    return tuple(
        etiqueta for attr, etiqueta in _CAMPOS_PERFIL_CLAVE
        if not getattr(contacto, attr, None)
    )


async def _publicar_perfil_incompleto(contacto) -> None:
    """Publica MiembroPerfilIncompleto si faltan campos clave. Nunca lanza."""
    faltantes = _campos_perfil_faltantes(contacto)
    if not faltantes:
        return
    try:
        nombre = " ".join(filter(None, [contacto.nombre, contacto.apellido1])).strip()
        await event_bus.publish(MiembroPerfilIncompleto(
            miembro_id=str(contacto.id),
            miembro_nombre=nombre or "(sin nombre)",
            agrupacion_id=str(contacto.agrupacion_id) if contacto.agrupacion_id else None,
            campos_faltantes=faltantes,
        ))
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Inputs
# ---------------------------------------------------------------------------

@strawberry.input
class MiembroCreateInput:
    nombre: str
    apellido1: str
    tipo_miembro_id: Optional[uuid.UUID] = None
    estado_id: Optional[uuid.UUID] = None
    apellido2: Optional[str] = None
    sexo: Optional[str] = None
    fecha_nacimiento: Optional[date] = None
    tipo_documento: Optional[str] = None
    numero_documento: Optional[str] = None
    pais_documento_id: Optional[uuid.UUID] = None
    pais_nacimiento_id: Optional[uuid.UUID] = None
    direccion: Optional[str] = None
    codigo_postal: Optional[str] = None
    localidad: Optional[str] = None
    provincia_id: Optional[uuid.UUID] = None
    pais_domicilio_id: Optional[uuid.UUID] = None
    telefono: Optional[str] = None
    telefono2: Optional[str] = None
    email: Optional[str] = None
    agrupacion_id: Optional[uuid.UUID] = None
    iban: Optional[str] = None
    swift_bic: Optional[str] = None
    referencia_pago: Optional[str] = None
    forma_pago_id: Optional[uuid.UUID] = None
    es_socio_honor: bool = False
    fecha_alta: Optional[date] = None
    fecha_baja: Optional[date] = None
    motivo_baja_id: Optional[uuid.UUID] = None
    motivo_baja_texto: Optional[str] = None
    observaciones: Optional[str] = None
    solicita_supresion_datos: bool = False
    fecha_solicitud_supresion: Optional[date] = None
    fecha_limite_retencion: Optional[date] = None
    datos_anonimizados: bool = False
    fecha_anonimizacion: Optional[date] = None
    activo: bool = True
    es_voluntario: bool = False
    disponibilidad: Optional[str] = None
    horas_disponibles_semana: Optional[int] = None
    profesion: Optional[str] = None
    nivel_estudios_id: Optional[uuid.UUID] = None
    motivo_reduccion_id: Optional[uuid.UUID] = None
    experiencia_voluntariado: Optional[str] = None
    intereses: Optional[str] = None
    observaciones_voluntariado: Optional[str] = None
    puede_conducir: bool = False
    vehiculo_propio: bool = False
    disponibilidad_viajar: bool = False


@strawberry.input
class MiembroUpdateInput:
    id: uuid.UUID
    nombre: Optional[str] = None
    apellido1: Optional[str] = None
    tipo_miembro_id: Optional[uuid.UUID] = None
    estado_id: Optional[uuid.UUID] = None
    apellido2: Optional[str] = None
    sexo: Optional[str] = None
    fecha_nacimiento: Optional[date] = None
    tipo_documento: Optional[str] = None
    numero_documento: Optional[str] = None
    pais_documento_id: Optional[uuid.UUID] = None
    pais_nacimiento_id: Optional[uuid.UUID] = None
    direccion: Optional[str] = None
    codigo_postal: Optional[str] = None
    localidad: Optional[str] = None
    provincia_id: Optional[uuid.UUID] = None
    pais_domicilio_id: Optional[uuid.UUID] = None
    telefono: Optional[str] = None
    telefono2: Optional[str] = None
    email: Optional[str] = None
    agrupacion_id: Optional[uuid.UUID] = None
    iban: Optional[str] = None
    swift_bic: Optional[str] = None
    referencia_pago: Optional[str] = None
    forma_pago_id: Optional[uuid.UUID] = None
    es_socio_honor: Optional[bool] = None
    fecha_alta: Optional[date] = None
    fecha_baja: Optional[date] = None
    motivo_baja_id: Optional[uuid.UUID] = None
    motivo_baja_texto: Optional[str] = None
    observaciones: Optional[str] = None
    solicita_supresion_datos: Optional[bool] = None
    fecha_solicitud_supresion: Optional[date] = None
    fecha_limite_retencion: Optional[date] = None
    datos_anonimizados: Optional[bool] = None
    fecha_anonimizacion: Optional[date] = None
    activo: Optional[bool] = None
    es_voluntario: Optional[bool] = None
    disponibilidad: Optional[str] = None
    horas_disponibles_semana: Optional[int] = None
    profesion: Optional[str] = None
    nivel_estudios_id: Optional[uuid.UUID] = None
    motivo_reduccion_id: Optional[uuid.UUID] = None
    experiencia_voluntariado: Optional[str] = None
    intereses: Optional[str] = None
    observaciones_voluntariado: Optional[str] = None
    puede_conducir: Optional[bool] = None
    vehiculo_propio: Optional[bool] = None
    disponibilidad_viajar: Optional[bool] = None


# ---------------------------------------------------------------------------
# Mixin de mutaciones
# ---------------------------------------------------------------------------

# Reparto del input plano a cada modelo del CRM (decisión: alta = Contacto +
# Vinculacion SOCIO + Socio + Participacion(MEMBRESIA)+Membresia [+ Voluntario]).
_CONTACTO_FIELDS = (
    'nombre', 'apellido1', 'apellido2', 'sexo', 'fecha_nacimiento', 'tipo_documento',
    'numero_documento', 'pais_documento_id', 'pais_nacimiento_id', 'direccion',
    'codigo_postal', 'localidad', 'provincia_id', 'pais_domicilio_id', 'telefono',
    'telefono2', 'email', 'agrupacion_id', 'profesion', 'nivel_estudios_id',
    'activo', 'solicita_supresion_datos', 'fecha_solicitud_supresion',
    'fecha_limite_retencion', 'datos_anonimizados', 'fecha_anonimizacion',
)
_SOCIO_FIELDS = (
    'iban', 'swift_bic', 'referencia_pago', 'forma_pago_id', 'motivo_reduccion_id',
    'motivo_baja_id', 'motivo_baja_texto',
)
# Dimensión ECONÓMICA del socio (competencia de Tesorería): editarla exige el
# permiso MEMBRESIA_MIEMBRO_EDITAR_DATOS_ECONOMICOS, aparte del permiso de edición registral. Los
# motivo_baja_* quedan fuera: son acto registral (situación), no económico.
_ECONOMIC_FIELDS = frozenset({
    'iban', 'swift_bic', 'referencia_pago', 'forma_pago_id', 'motivo_reduccion_id',
})
_VOLUNTARIO_FIELDS = (
    'disponibilidad', 'horas_disponibles_semana', 'profesion', 'nivel_estudios_id',
    'experiencia_voluntariado', 'intereses', 'observaciones_voluntariado',
    'puede_conducir', 'vehiculo_propio', 'disponibilidad_viajar',
)


async def _tipo_vinc_id(session, codigo: str):
    return await session.scalar(
        select(TipoVinculacion.id).where(TipoVinculacion.codigo == codigo)
    )


async def _vinc_socio(session, contacto_id: uuid.UUID):
    """Vinculación SOCIO (cualquier estado) más reciente del contacto, o None."""
    return (await session.execute(
        select(Vinculacion)
        .join(TipoVinculacion, Vinculacion.tipo_vinculacion_id == TipoVinculacion.id)
        .where(TipoVinculacion.codigo == "SOCIO", Vinculacion.contacto_id == contacto_id)
        .order_by(Vinculacion.fecha_inicio.desc())
    )).scalars().first()


async def _vinc_voluntario(session, contacto_id: uuid.UUID):
    return (await session.execute(
        select(Vinculacion)
        .join(TipoVinculacion, Vinculacion.tipo_vinculacion_id == TipoVinculacion.id)
        .where(TipoVinculacion.codigo == "VOLUNTARIO", Vinculacion.contacto_id == contacto_id)
        .order_by(Vinculacion.fecha_inicio.desc())
    )).scalars().first()


async def _membresia_de(session, contacto_id: uuid.UUID):
    return (await session.execute(
        select(Membresia)
        .join(Participacion, Membresia.participacion_id == Participacion.id)
        .where(Participacion.contacto_id == contacto_id, Participacion.tipo == "MEMBRESIA")
        .order_by(Participacion.fecha.desc())
    )).scalars().first()


async def _alta_socio(session, data) -> Contacto:
    """Crea Contacto + Participacion(MEMBRESIA)+Membresia + Vinculacion(SOCIO)+Socio
    y, si `es_voluntario`, Vinculacion(VOLUNTARIO)+Voluntario. Devuelve el Contacto."""
    contacto = Contacto(tipo="PERSONA_FISICA",
                        **{f: getattr(data, f, None) for f in _CONTACTO_FIELDS})
    session.add(contacto)
    await session.flush()

    # Acto de alta: Participacion(MEMBRESIA) + Membresia (lleva el tipo_miembro)
    part = Participacion(contacto_id=contacto.id, tipo="MEMBRESIA")
    session.add(part)
    await session.flush()
    session.add(Membresia(participacion_id=part.id,
                          tipo_miembro_id=getattr(data, "tipo_miembro_id", None)))

    # Vinculación de socio + satélite económico
    baja = getattr(data, "fecha_baja", None)
    socio_vinc = Vinculacion(
        contacto_id=contacto.id,
        tipo_vinculacion_id=await _tipo_vinc_id(session, "SOCIO"),
        fecha_inicio=getattr(data, "fecha_alta", None) or date.today(),
        fecha_fin=baja,
        estado="baja" if baja else "activa",
        agrupacion_id=getattr(data, "agrupacion_id", None),
    )
    session.add(socio_vinc)
    await session.flush()
    session.add(Socio(
        vinculacion_id=socio_vinc.id,
        es_honor=bool(getattr(data, "es_socio_honor", False)),
        estado_socio="baja" if baja else "activo",
        **{f: getattr(data, f, None) for f in _SOCIO_FIELDS},
    ))

    # Voluntario opcional (otra vinculación + satélite)
    if getattr(data, "es_voluntario", False):
        vol_vinc = Vinculacion(
            contacto_id=contacto.id,
            tipo_vinculacion_id=await _tipo_vinc_id(session, "VOLUNTARIO"),
            fecha_inicio=date.today(), estado="activa",
            agrupacion_id=getattr(data, "agrupacion_id", None),
        )
        session.add(vol_vinc)
        await session.flush()
        session.add(Voluntario(vinculacion_id=vol_vinc.id,
                               **{f: getattr(data, f, None) for f in _VOLUNTARIO_FIELDS}))

    return contacto


async def _fetch_miembro(session, contacto_id: uuid.UUID):
    """Recarga el Contacto (identidad viva) con sus relaciones selectin."""
    stmt = select(Contacto).where(Contacto.id == contacto_id)
    result = await session.execute(stmt)
    return result.scalar_one()


@strawberry.input
class PerfilVoluntarioInput:
    """Campos de voluntariado editables por delegación (no toca el resto del socio)."""
    miembro_id: uuid.UUID
    es_voluntario: Optional[bool] = None
    disponibilidad: Optional[str] = None
    horas_disponibles_semana: Optional[int] = None
    profesion: Optional[str] = None
    nivel_estudios_id: Optional[uuid.UUID] = None
    intereses: Optional[str] = None
    experiencia_voluntariado: Optional[str] = None
    observaciones_voluntariado: Optional[str] = None
    puede_conducir: Optional[bool] = None
    vehiculo_propio: Optional[bool] = None
    disponibilidad_viajar: Optional[bool] = None


@strawberry.type
class MembresiaResolverMutation:

    @strawberry.mutation(permission_classes=[RequireTransaction("MEMBRESIA_VOLUNTARIO_GESTIONAR")])
    async def gestionar_perfil_voluntario(
        self,
        info: strawberry.Info,
        data: PerfilVoluntarioInput,
    ) -> 'ContactoType':
        """Edita el perfil de voluntario (disponibilidad, profesión, intereses…) por delegación.

        El voluntariado es una Vinculacion(VOLUNTARIO) con satélite Voluntario: se crea
        si `es_voluntario=True` y no existía, o se da de baja si `es_voluntario=False`.
        profesion/nivel_estudios se guardan también en el Contacto. Aplica el guard de
        ámbito territorial.
        """
        from app.modules.acceso.services.ambito_territorial import assert_miembro_en_ambito
        session = info.context.session
        user = info.context.user
        await assert_miembro_en_ambito(session, user.id, data.miembro_id)

        contacto = await _fetch_miembro(session, data.miembro_id)
        for field in ("profesion", "nivel_estudios_id"):
            val = getattr(data, field, None)
            if val is not None:
                setattr(contacto, field, val)

        vinc = await _vinc_voluntario(session, contacto.id)
        if data.es_voluntario is False:
            if vinc:
                vinc.estado = "baja"
                vinc.fecha_fin = date.today()
        else:
            if not vinc:
                vinc = Vinculacion(
                    contacto_id=contacto.id,
                    tipo_vinculacion_id=await _tipo_vinc_id(session, "VOLUNTARIO"),
                    fecha_inicio=date.today(), estado="activa",
                    agrupacion_id=contacto.agrupacion_id,
                )
                session.add(vinc)
                await session.flush()
            vol = vinc.voluntario
            if not vol:
                vol = Voluntario(vinculacion_id=vinc.id)
                session.add(vol)
            for field in _VOLUNTARIO_FIELDS:
                val = getattr(data, field, None)
                if val is not None:
                    setattr(vol, field, val)
        await session.commit()
        return await _fetch_miembro(session, contacto.id)

    @strawberry.mutation(permission_classes=[RequireTransaction("MEMBRESIA_VOLUNTARIO_GESTIONAR")])
    async def asignar_habilidad_voluntario(
        self,
        info: strawberry.Info,
        miembro_id: uuid.UUID,
        habilidad_id: uuid.UUID,
        nivel_id: Optional[uuid.UUID] = None,
    ) -> bool:
        """Asigna (o actualiza el nivel de) una habilidad a un socio, por delegación.
        Con guard de ámbito territorial."""
        from app.modules.acceso.services.ambito_territorial import assert_miembro_en_ambito
        from app.modules.membresia.models.habilidad import MiembroHabilidad
        session = info.context.session
        await assert_miembro_en_ambito(session, info.context.user.id, miembro_id)
        # Las habilidades cuelgan de la extensión Voluntario: traducir contacto → voluntario.
        vinc = await _vinc_voluntario(session, miembro_id)
        if not vinc or not vinc.voluntario:
            raise ValueError("El contacto no tiene una vinculación de voluntario; "
                             "las habilidades pertenecen al perfil de voluntario.")
        voluntario_id = vinc.voluntario.id
        existente = (await session.execute(
            select(MiembroHabilidad).where(
                MiembroHabilidad.voluntario_id == voluntario_id,
                MiembroHabilidad.habilidad_id == habilidad_id,
            )
        )).scalar_one_or_none()
        if existente:
            existente.nivel_id = nivel_id
        else:
            session.add(MiembroHabilidad(
                voluntario_id=voluntario_id, habilidad_id=habilidad_id, nivel_id=nivel_id,
            ))
        await session.commit()
        return True

    @strawberry.mutation(permission_classes=[RequireTransaction("MEMBRESIA_VOLUNTARIO_GESTIONAR")])
    async def quitar_habilidad_voluntario(
        self,
        info: strawberry.Info,
        miembro_id: uuid.UUID,
        habilidad_id: uuid.UUID,
    ) -> bool:
        """Quita una habilidad de un socio, por delegación. Con guard de ámbito territorial."""
        from app.modules.acceso.services.ambito_territorial import assert_miembro_en_ambito
        from app.modules.membresia.models.habilidad import MiembroHabilidad
        session = info.context.session
        await assert_miembro_en_ambito(session, info.context.user.id, miembro_id)
        # Las habilidades cuelgan de la extensión Voluntario: traducir contacto → voluntario.
        vinc = await _vinc_voluntario(session, miembro_id)
        if not vinc or not vinc.voluntario:
            return False
        existente = (await session.execute(
            select(MiembroHabilidad).where(
                MiembroHabilidad.voluntario_id == vinc.voluntario.id,
                MiembroHabilidad.habilidad_id == habilidad_id,
            )
        )).scalar_one_or_none()
        if existente:
            await session.delete(existente)
            await session.commit()
        return True

    @strawberry.mutation(permission_classes=[RequireTransaction("MEMBRESIA_MIEMBRO_CREAR")])
    async def crear_miembro(
        self,
        info: strawberry.Info,
        data: MiembroCreateInput,
    ) -> 'ContactoType':
        session = info.context.session

        contacto = await _alta_socio(session, data)
        await session.commit()

        await _publicar_perfil_incompleto(contacto)

        return await _fetch_miembro(session, contacto.id)

    @strawberry.mutation(permission_classes=[RequireTransaction("MEMBRESIA_MIEMBRO_CREAR")])
    async def crear_miembro_con_acceso(
        self,
        info: strawberry.Info,
        data: MiembroCreateInput,
        email: str,
        password: str,
        tipo_vinculacion_id: Optional[uuid.UUID] = None,
        activo_usuario: bool = True,
    ) -> 'ContactoType':
        """Crea un socio (Contacto + satélites) y su usuario de acceso en una única
        transacción atómica."""
        session = info.context.session

        contacto = await _alta_socio(session, data)

        svc = AccesoService(session)
        await svc.crear_usuario(
            email=email,
            password=password,
            activo=activo_usuario,
            contacto_id=contacto.id,
            tipo_vinculacion_id=tipo_vinculacion_id,
        )

        await session.commit()

        await _publicar_perfil_incompleto(contacto)

        return await _fetch_miembro(session, contacto.id)

    @strawberry.mutation(permission_classes=[RequireTransaction("MEMBRESIA_MIEMBRO_EDITAR")])
    async def actualizar_miembro(
        self,
        info: strawberry.Info,
        data: MiembroUpdateInput,
    ) -> 'ContactoType':
        session = info.context.session

        miembro = await _fetch_miembro(session, data.id)
        faltantes_antes = _campos_perfil_faltantes(miembro)

        # Enforcement por dimensión: los datos económicos del socio son competencia
        # de Tesorería. Si la edición los toca, exige MEMBRESIA_MIEMBRO_EDITAR_DATOS_ECONOMICOS (además del
        # permiso registral de la mutación). Se comprueba antes de aplicar nada.
        if any(getattr(data, f, None) is not None for f in _ECONOMIC_FIELDS):
            if not await info.context.check_permission("MEMBRESIA_MIEMBRO_EDITAR_DATOS_ECONOMICOS"):
                raise ValueError(
                    "No tienes permiso para editar los datos económicos del socio "
                    "(IBAN, forma de pago, reducción de cuota): es competencia de Tesorería."
                )

        # Datos de identidad -> Contacto (None = no tocar)
        for field in _CONTACTO_FIELDS:
            val = getattr(data, field, None)
            if val is not None:
                setattr(miembro, field, val)

        # tipo de socio -> Membresía
        if getattr(data, "tipo_miembro_id", None) is not None:
            memb = await _membresia_de(session, miembro.id)
            if memb:
                memb.tipo_miembro_id = data.tipo_miembro_id

        # datos económicos / baja -> satélite Socio
        vinc = await _vinc_socio(session, miembro.id)
        socio = vinc.socio if vinc else None
        if socio:
            for field in _SOCIO_FIELDS:
                val = getattr(data, field, None)
                if val is not None:
                    setattr(socio, field, val)
            if getattr(data, "es_socio_honor", None) is not None:
                socio.es_honor = data.es_socio_honor

        await session.commit()

        # Notificar solo si la edición ha cambiado el conjunto de campos
        # faltantes y siguen quedando huecos. Si los huecos son los mismos
        # que antes, evitamos el spam por cada edición del coordinador.
        faltantes_despues = _campos_perfil_faltantes(miembro)
        if faltantes_despues and faltantes_despues != faltantes_antes:
            await _publicar_perfil_incompleto(miembro)

        return await _fetch_miembro(session, miembro.id)

    @strawberry.mutation(permission_classes=[RequireTransaction("MEMBRESIA_MIEMBRO_EDITAR")])
    async def anonimizar_miembro(
        self,
        info: strawberry.Info,
        miembro_id: uuid.UUID,
    ) -> 'ContactoType':
        """RGPD — anonimiza de forma irreversible los datos personales del
        miembro. El registro se conserva (para estadística e histórico) pero
        despojado de toda información identificativa. Solo se permite sobre
        miembros dados de baja.
        """
        session = info.context.session
        miembro = await _fetch_miembro(session, miembro_id)

        if miembro.datos_anonimizados:
            raise ValueError("Los datos de este contacto ya están anonimizados.")
        # La baja vive en la vinculación de socio (fecha_fin) / satélite Socio.
        vinc = await _vinc_socio(session, miembro.id)
        if not (vinc and vinc.fecha_fin is not None):
            raise ValueError(
                "Solo pueden anonimizarse los datos de un socio dado de baja."
            )

        # Despojar al Contacto de toda información personal identificativa.
        miembro.nombre = "Anonimizado"
        miembro.apellido1 = "Anonimizado"
        miembro.apellido2 = None
        miembro.fecha_nacimiento = None
        miembro.sexo = None
        miembro.tipo_documento = None
        miembro.numero_documento = None
        miembro.direccion = None
        miembro.codigo_postal = None
        miembro.localidad = None
        miembro.telefono = None
        miembro.telefono2 = None
        miembro.email = None
        miembro.foto_url = None
        miembro.observaciones = None

        # Datos bancarios del socio (en el satélite)
        socio = vinc.socio
        if socio:
            socio.iban = None
            socio.swift_bic = None
            socio.referencia_pago = None

        miembro.datos_anonimizados = True
        miembro.fecha_anonimizacion = date.today()

        await session.commit()
        return await _fetch_miembro(session, miembro_id)

    @strawberry.mutation(permission_classes=[RequireTransaction("MEMBRESIA_MIEMBRO_EXPORTAR")])
    async def exportar_miembros_xlsx(
        self,
        info: strawberry.Info,
        ids: List[uuid.UUID],
    ) -> str:
        """Exporta a XLSX los miembros indicados (los visibles con los filtros
        aplicados en el listado). Devuelve el contenido del fichero en base64.
        """
        import base64
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        session = info.context.session
        if not ids:
            raise ValueError("No hay miembros que exportar.")

        result = await session.execute(select(Contacto).where(Contacto.id.in_(ids)))
        contactos = list(result.scalars().all())
        contactos.sort(key=lambda c: (
            (c.apellido1 or '').lower(),
            (c.apellido2 or '').lower(),
            (c.nombre or '').lower(),
        ))

        # Nombres de agrupación (Contacto solo guarda el id)
        from app.modules.core.geografico.direccion import UnidadOrganizativa
        agr_ids = {c.agrupacion_id for c in contactos if c.agrupacion_id}
        agr_nombres: dict = {}
        if agr_ids:
            ra = await session.execute(
                select(UnidadOrganizativa.id, UnidadOrganizativa.nombre)
                .where(UnidadOrganizativa.id.in_(agr_ids))
            )
            agr_nombres = {i: n for (i, n) in ra.all()}

        wb = Workbook()
        ws = wb.active
        ws.title = "Socios"
        cabecera = [
            "Nombre", "Primer apellido", "Segundo apellido", "Tipo", "Situación",
            "Email", "Teléfono", "Agrupación", "Localidad", "Fecha de alta", "Fecha de baja",
        ]
        ws.append(cabecera)
        for celda in ws[1]:
            celda.font = Font(bold=True)

        for c in contactos:
            memb = await _membresia_de(session, c.id)
            vinc = await _vinc_socio(session, c.id)
            socio = vinc.socio if vinc else None
            ws.append([
                c.nombre or '',
                c.apellido1 or '',
                c.apellido2 or '',
                memb.tipo_miembro.nombre if (memb and memb.tipo_miembro) else '',
                (socio.estado_socio if socio else '') or '',
                c.email or '',
                c.telefono or '',
                agr_nombres.get(c.agrupacion_id, ''),
                c.localidad or '',
                vinc.fecha_inicio.isoformat() if (vinc and vinc.fecha_inicio) else '',
                vinc.fecha_fin.isoformat() if (vinc and vinc.fecha_fin) else '',
            ])

        anchos = [16, 16, 16, 14, 14, 30, 14, 26, 20, 13, 13]
        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        return base64.b64encode(buf.getvalue()).decode()


@strawberry.type(name="VoluntarioAmbito")
class VoluntarioAmbitoType:
    """Voluntario (subconjunto de Miembro) para el listado con scoping de ámbito.

    Vista de listado; no confundir con la entidad `Voluntario` (satélite de
    Vinculacion del modelo CRM), que posee el nombre GraphQL `Voluntario`.
    """
    id: uuid.UUID
    nombre: str
    apellido1: str
    apellido2: Optional[str]
    email: Optional[str]
    telefono: Optional[str]
    disponibilidad: Optional[str]
    horas_disponibles_semana: Optional[int]
    profesion: Optional[str]
    nivel_estudios_id: Optional[uuid.UUID]
    intereses: Optional[str]
    puede_conducir: bool
    vehiculo_propio: bool
    disponibilidad_viajar: bool
    activo: bool
    fecha_alta: Optional[date]


@strawberry.type(name="ContactoCondiciones")
class ContactoCondicionesType:
    """Condiciones DERIVADAS de un contacto (no son vinculaciones): se calculan a
    partir de sus registros. Alimentan los badges de la ficha del contacto."""
    es_participante: bool
    es_firmante: bool
    es_donante: bool
    n_firmas: int
    n_participaciones: int
    n_donaciones: int


@strawberry.type(name="ContactoCondicionesItem")
class ContactoCondicionesItemType:
    """Condiciones derivadas de un contacto (versión ligera para listados)."""
    contacto_id: uuid.UUID
    es_participante: bool
    es_firmante: bool
    es_donante: bool


@strawberry.type
class MembresiaQuery:
    @strawberry.field(permission_classes=[RequireAuthenticated])
    async def condiciones_contacto(
        self, info: strawberry.Info, contacto_id: uuid.UUID
    ) -> ContactoCondicionesType:
        """Condiciones derivadas (firmante/participante/donante) de un contacto,
        calculadas de sus registros (Participacion/FirmaCampania/Donacion)."""
        from sqlalchemy import func

        from app.modules.actividades.models.campana import FirmaCampania
        from app.modules.economico.models.donaciones import Donacion

        session = info.context.session
        n_firmas = await session.scalar(
            select(func.count(FirmaCampania.id)).where(
                FirmaCampania.contacto_id == contacto_id,
                FirmaCampania.eliminado.is_(False),
            )
        ) or 0
        n_part = await session.scalar(
            select(func.count(Participacion.id)).where(
                Participacion.contacto_id == contacto_id,
                Participacion.tipo.in_(("FIRMA", "ASISTENCIA")),
                Participacion.eliminado.is_(False),
            )
        ) or 0
        n_don = await session.scalar(
            select(func.count(Donacion.id)).where(
                Donacion.contacto_id == contacto_id,
                Donacion.eliminado.is_(False),
            )
        ) or 0
        return ContactoCondicionesType(
            es_participante=n_part > 0,
            es_firmante=n_firmas > 0,
            es_donante=n_don > 0,
            n_firmas=int(n_firmas),
            n_participaciones=int(n_part),
            n_donaciones=int(n_don),
        )

    @strawberry.field(permission_classes=[RequireAuthenticated])
    async def condiciones_contactos(
        self, info: strawberry.Info, contacto_ids: List[uuid.UUID]
    ) -> List[ContactoCondicionesItemType]:
        """Condiciones derivadas de VARIOS contactos (batch, para listados). Evita
        el N+1: 3 consultas agrupadas por contacto para firmas, participaciones y
        donaciones."""
        from sqlalchemy import func

        from app.modules.actividades.models.campana import FirmaCampania
        from app.modules.economico.models.donaciones import Donacion

        session = info.context.session
        ids = list({cid for cid in contacto_ids})
        if not ids:
            return []

        async def _contar(col_modelo, extra=None):
            stmt = select(col_modelo, func.count()).where(col_modelo.in_(ids))
            if extra is not None:
                stmt = stmt.where(extra)
            stmt = stmt.group_by(col_modelo)
            return {row[0]: row[1] for row in (await session.execute(stmt)).all()}

        firmas = await _contar(FirmaCampania.contacto_id, FirmaCampania.eliminado.is_(False))
        parts = await _contar(
            Participacion.contacto_id,
            (Participacion.tipo.in_(("FIRMA", "ASISTENCIA"))) & (Participacion.eliminado.is_(False)),
        )
        dons = await _contar(Donacion.contacto_id, Donacion.eliminado.is_(False))

        return [
            ContactoCondicionesItemType(
                contacto_id=cid,
                es_firmante=firmas.get(cid, 0) > 0,
                es_participante=parts.get(cid, 0) > 0,
                es_donante=dons.get(cid, 0) > 0,
            )
            for cid in ids
        ]

    @strawberry.field(permission_classes=[RequireTransaction("MEMBRESIA_VOLUNTARIO_LISTAR")])
    async def voluntarios_en_ambito(self, info: strawberry.Info) -> List[VoluntarioAmbitoType]:
        """Voluntarios visibles según el ámbito territorial del usuario (Fase 2).

        - Rol general (presidencia / rol en unidad raíz) → todos los voluntarios.
        - Rol territorial → voluntarios de su agrupación + descendientes.
        - Sin ámbito territorial → lista vacía por esta vía.
        (El scoping por campaña de COORDINADOR_CAMPANA se resolverá aparte.)
        """
        from app.modules.acceso.services.ambito_territorial import (
            agrupaciones_en_ambito, miembros_de_campanias_coordinadas,
        )
        session = info.context.session
        user = info.context.user
        ambito = await agrupaciones_en_ambito(session, user.id) if user else set()

        # Voluntario = Vinculacion(VOLUNTARIO) activa + satélite Voluntario sobre Contacto.
        q = (
            select(Contacto, Voluntario, Vinculacion)
            .join(Vinculacion, Vinculacion.contacto_id == Contacto.id)
            .join(TipoVinculacion, Vinculacion.tipo_vinculacion_id == TipoVinculacion.id)
            .join(Voluntario, Voluntario.vinculacion_id == Vinculacion.id)
            .where(
                TipoVinculacion.codigo == "VOLUNTARIO",
                Vinculacion.estado == "activa",
                Contacto.eliminado == False,    # noqa: E712
            )
        )
        if ambito is not None:              # None = global (sin filtro)
            conds = []
            if ambito:
                conds.append(Contacto.agrupacion_id.in_(ambito))
            camp_ids = await miembros_de_campanias_coordinadas(session, user.id) if user else set()
            if camp_ids:
                conds.append(Contacto.id.in_(camp_ids))
            if not conds:
                return []
            q = q.where(or_(*conds))

        r = await session.execute(q.order_by(Contacto.apellido1, Contacto.nombre))
        return [
            VoluntarioAmbitoType(
                id=c.id, nombre=c.nombre, apellido1=c.apellido1 or "", apellido2=c.apellido2,
                email=c.email, telefono=c.telefono, disponibilidad=v.disponibilidad,
                horas_disponibles_semana=v.horas_disponibles_semana, profesion=v.profesion,
                nivel_estudios_id=v.nivel_estudios_id, intereses=v.intereses,
                puede_conducir=v.puede_conducir, vehiculo_propio=v.vehiculo_propio,
                disponibilidad_viajar=v.disponibilidad_viajar,
                activo=c.activo, fecha_alta=vinc.fecha_inicio,
            )
            for (c, v, vinc) in r.all()
        ]
